"""
Parser for the monthly oil price-report workbooks, both categories.

Each workbook has one INDEX sheet (skipped) plus a set of product sheets.
Every product sheet is a daily price grid anchored on a "DATE" label: the
column under that label holds dates, every column to its right is a distinct
price series whose name is spread across the header rows above and below the
DATE row. This module flattens each sheet into a long/tall list of ParsedRow.

The two report formats differ in ways this parser handles generically:

  NON_EDIBLE  DATE label in column A; header rows sit above it; data starts on
              the row immediately after.
  EDIBLE      DATE label is not always in column A (CASTOR starts at column F);
              qualifier rows ("+GST", the trading centre) sit *below* the DATE
              row; prices are often quoted as a range ("6600-6800"); sheets
              carry side-by-side sub-tables, each with its own DATE column and
              a MONTH/CLOSE/CHANGE futures block; and a sheet may re-declare
              its contract-month header partway down on a row that still has a
              valid date (CHINA does this when contracts roll).

Values are classified into: a number, a low-high range, a text label (a
futures MONTH column), or a status token ("NA", "SUNDAY", "CLOSE", ...).
"""
from __future__ import annotations

import datetime
import re
from dataclasses import dataclass, field

import openpyxl

SKIP_SHEETS = {"INDEX"}

STATUS_TOKENS = {"NA", "SUNDAY", "MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY",
                 "FRIDAY", "SATURDAY", "CLOSE", "CLOSED", "CLS", "HOLIDAY",
                 "HOIDAY", "NSLR", "NQ"}

# Contract labels used by the futures blocks. A data row made up entirely of
# these is a header the sheet re-declared mid-grid, not prices.
MONTH_TOKENS = {"JAN", "FEB", "MAR", "APR", "MAY", "JUN", "JUNE", "JUL", "JULY",
                "AUG", "SEP", "SEPT", "OCT", "NOV", "DEC",
                "ND", "JFM", "OND", "AMJ", "JAS", "FMA", "MJJ", "ASO", "NDJ"}

# "6600-6800", "1137.5 - 1150" (hyphen or en-dash)
_RANGE_RE = re.compile(r"^(\d+(?:\.\d+)?)\s*[-–]\s*(\d+(?:\.\d+)?)$")

_MAX_COL_SCAN = 250      # SOY OIL REF reaches column 92, RICEBRAN 100
_MAX_HEADER_SCAN = 15    # deepest DATE label seen is row 3
_MAX_DATE_COL_SCAN = 30  # CASTOR's DATE label sits at column 6
_MAX_COL_GAP = 25        # stop scanning after this many empty header columns

# Value kinds
KIND_NUMBER = "NUMBER"
KIND_RANGE = "RANGE"
KIND_TEXT = "TEXT"
KIND_STATUS = "STATUS"


@dataclass
class ParsedRow:
    sheet_name: str
    series_name: str
    column_index: int
    price_date: datetime.date
    raw_value: str
    price_value: float | None       # midpoint for a range; the number otherwise
    price_low: float | None
    price_high: float | None
    text_value: str | None          # futures MONTH label
    status_flag: str | None
    kind: str = KIND_NUMBER


@dataclass
class SheetParseResult:
    sheet_name: str
    series: list[str] = field(default_factory=list)
    rows: list[ParsedRow] = field(default_factory=list)
    date_range: tuple[datetime.date, datetime.date] | None = None
    warnings: list[str] = field(default_factory=list)
    # Populated by pivot_sheet_wide(): wide-frame column -> "FLOAT" | "VARCHAR"
    wide_column_types: dict = field(default_factory=dict)


@dataclass
class WorkbookParseResult:
    file_name: str
    sheets: list[SheetParseResult] = field(default_factory=list)

    @property
    def total_rows(self) -> int:
        return sum(len(s.rows) for s in self.sheets)

    @property
    def all_series(self) -> list[tuple[str, str]]:
        return [(s.sheet_name, series) for s in self.sheets for series in s.series]

    def to_dataframe(self):
        import pandas as pd
        records = []
        for sheet in self.sheets:
            for r in sheet.rows:
                records.append({
                    "SHEET_NAME": r.sheet_name,
                    "SERIES_NAME": r.series_name,
                    "COLUMN_INDEX": r.column_index,
                    "PRICE_DATE": r.price_date,
                    "RAW_VALUE": r.raw_value,
                    "PRICE_VALUE": r.price_value,
                    "PRICE_LOW": r.price_low,
                    "PRICE_HIGH": r.price_high,
                    "TEXT_VALUE": r.text_value,
                    "STATUS_FLAG": r.status_flag,
                    "KIND": r.kind,
                })
        return pd.DataFrame.from_records(records)

    def to_wide_frames(self):
        """One wide, Excel-shaped DataFrame per sheet: PRICE_DATE, DAY_TYPE,
        then LOW/HIGH columns per price series. See pivot_sheet_wide()."""
        return {sheet.sheet_name: pivot_sheet_wide(sheet) for sheet in self.sheets}


# --------------------------------------------------------------------------- #
# cell helpers
# --------------------------------------------------------------------------- #
def _build_merge_map(ws):
    """Map every cell coordinate covered by a merge to the anchor cell's value."""
    merge_map = {}
    for merged_range in ws.merged_cells.ranges:
        anchor_value = ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                merge_map[(row, col)] = anchor_value
    return merge_map


def _cell_value(ws, merge_map, row, col):
    if (row, col) in merge_map:
        return merge_map[(row, col)]
    return ws.cell(row=row, column=col).value


def _clean(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def _is_letterhead_noise(text: str) -> bool:
    """Company letterhead/address blocks sometimes live in a merged cell that
    spans real data columns; they must not pollute the series name."""
    if "@" in text or "FAX" in text.upper():
        return True
    if len(text) > 60:
        return True
    return False


def _is_date(v) -> bool:
    return isinstance(v, (datetime.datetime, datetime.date))


def _as_date(v):
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    return None


# --------------------------------------------------------------------------- #
# sheet geometry
# --------------------------------------------------------------------------- #
def _find_date_anchor(ws, merge_map) -> tuple[int, int] | None:
    """Locate the 'DATE' label. Returns (row, col), or None if absent.

    Scans left-to-right within each row so the leftmost block wins on sheets
    that carry several side-by-side sub-tables.
    """
    for row in range(1, _MAX_HEADER_SCAN + 1):
        for col in range(1, _MAX_DATE_COL_SCAN + 1):
            if _clean(_cell_value(ws, merge_map, row, col)).upper() == "DATE":
                return row, col
    # Fall back to a looser match ("DATE" embedded in a longer label).
    for row in range(1, _MAX_HEADER_SCAN + 1):
        for col in range(1, _MAX_DATE_COL_SCAN + 1):
            if "DATE" in _clean(_cell_value(ws, merge_map, row, col)).upper():
                return row, col
    return None


def _find_data_start(ws, date_row: int, date_col: int, max_row: int) -> int | None:
    """First row at/after the DATE row whose date column holds a real date.

    Rows between the DATE label and this one are qualifier headers ("+GST",
    the trading centre) that belong in the series name.
    """
    for row in range(date_row + 1, max_row + 1):
        if _is_date(ws.cell(row=row, column=date_col).value):
            return row
    return None


def _effective_max_row(ws, date_row: int) -> int:
    """openpyxl over-reports dimensions on sheets with phantom formatted rows
    (PALM OIL 1 and RICEBRAN claim ~65,000 rows); cap the scan to a sane
    window past the last plausible date."""
    real_max_row = ws.max_row
    if ws.title == "SPENT" or real_max_row > 5000:
        return date_row + 400
    return real_max_row


def _find_used_columns(ws, header_rows: range, date_col: int) -> int:
    """Rightmost column carrying header text.

    ws.max_column is unusable on some sheets (CASTOR reports 16,369), so scan
    the header band instead, stopping once the band has been empty for
    _MAX_COL_GAP columns. The widest real gap between populated header columns
    in the sample reports is 9 (DDGS), so the gap rule keeps every genuine
    block while ignoring stray text stranded far to the right.
    """
    max_col = date_col
    for col in range(date_col, _MAX_COL_SCAN + 1):
        if any(ws.cell(row=row, column=col).value not in (None, "") for row in header_rows):
            max_col = col
        elif col - max_col >= _MAX_COL_GAP:
            break
    return max_col


# --------------------------------------------------------------------------- #
# value classification
# --------------------------------------------------------------------------- #
def classify_value(value):
    """Classify one cell into (kind, low, high, text, status).

    A bare number is a degenerate range, so low == high; that keeps every
    numeric series two uniform FLOAT columns wide.
    """
    if isinstance(value, bool):
        return KIND_STATUS, None, None, None, str(value).upper()
    if isinstance(value, (int, float)):
        v = float(value)
        return KIND_NUMBER, v, v, None, None

    raw = _clean(value)
    token = raw.upper()

    m = _RANGE_RE.match(raw)
    if m:
        low, high = float(m.group(1)), float(m.group(2))
        if low > high:
            low, high = high, low
        return KIND_RANGE, low, high, None, None

    try:
        v = float(raw)
        return KIND_NUMBER, v, v, None, None
    except ValueError:
        pass

    if token in MONTH_TOKENS:
        return KIND_TEXT, None, None, token, None
    return KIND_STATUS, None, None, None, token or None


def _is_header_redeclaration(cells: dict) -> bool:
    """True when a dated row carries only contract labels.

    CHINA re-prints its MONTH header at row 17 when contracts roll; the row has
    a valid date, so without this it would load "SEP"/"NOV" as prices.
    """
    kinds = [classify_value(v)[0] for v in cells.values() if v not in (None, "")]
    if len(kinds) < 3:
        return False
    return all(k == KIND_TEXT for k in kinds)


# --------------------------------------------------------------------------- #
# parsing
# --------------------------------------------------------------------------- #
def parse_sheet(ws) -> SheetParseResult:
    result = SheetParseResult(sheet_name=ws.title)
    merge_map = _build_merge_map(ws)

    anchor = _find_date_anchor(ws, merge_map)
    if anchor is None:
        result.warnings.append("Could not locate a 'DATE' header row; sheet skipped.")
        return result
    date_row, date_col = anchor

    max_row = _effective_max_row(ws, date_row)
    data_start = _find_data_start(ws, date_row, date_col, max_row)
    if data_start is None:
        result.warnings.append("No dated rows found below the 'DATE' header; sheet skipped.")
        return result

    header_rows = range(1, data_start)
    used_max_col = _find_used_columns(ws, header_rows, date_col)

    # --- series names, from every header row above and below the DATE label ---
    series_by_col: dict[int, str] = {}
    for col in range(date_col + 1, used_max_col + 1):
        parts = []
        for row in header_rows:
            v = _clean(_cell_value(ws, merge_map, row, col))
            if v and v not in parts and not _is_letterhead_noise(v):
                parts.append(v)
        name = " | ".join(parts)
        if name:
            series_by_col[col] = name

    # --- drop the DATE columns of side-by-side sub-tables -------------------
    extra_date_cols = set()
    for col, name in series_by_col.items():
        if any(p.strip().upper() == "DATE" for p in name.split("|")):
            extra_date_cols.add(col)
            continue
        dated = sum(
            1 for row in range(data_start, min(data_start + 20, max_row + 1))
            if _is_date(ws.cell(row=row, column=col).value)
        )
        if dated >= 10:
            extra_date_cols.add(col)
    for col in extra_date_cols:
        series_by_col.pop(col, None)
    if extra_date_cols:
        result.warnings.append(
            f"Ignored {len(extra_date_cols)} duplicate DATE column(s) from side-by-side blocks."
        )

    result.series = list(series_by_col.values())
    if not series_by_col:
        result.warnings.append("No price series found to the right of the DATE column.")
        return result

    # --- collect the grid ---------------------------------------------------
    grid: dict[datetime.date, dict[int, object]] = {}
    min_d = max_d = None
    redeclared = 0
    consecutive_blank = 0
    row = data_start
    while row <= max_row:
        price_date = _as_date(ws.cell(row=row, column=date_col).value)
        if price_date is None:
            row_has_any_value = any(
                ws.cell(row=row, column=c).value not in (None, "")
                for c in range(date_col, used_max_col + 1)
            )
            consecutive_blank = 0 if row_has_any_value else consecutive_blank + 1
            if consecutive_blank >= 5:
                break
            row += 1
            continue

        consecutive_blank = 0
        cells = {col: ws.cell(row=row, column=col).value for col in series_by_col}
        if _is_header_redeclaration(cells):
            redeclared += 1
            row += 1
            continue

        min_d = price_date if min_d is None else min(min_d, price_date)
        max_d = price_date if max_d is None else max(max_d, price_date)
        # A repeated date (rare) merges into the row already collected.
        grid.setdefault(price_date, {}).update(
            {c: v for c, v in cells.items() if v not in (None, "")}
        )
        row += 1

    if redeclared:
        result.warnings.append(
            f"Skipped {redeclared} row(s) that re-declare the contract-month header."
        )

    # --- decide which columns are text (futures MONTH) labels ---------------
    text_cols = set()
    for col in series_by_col:
        kinds = [
            classify_value(cells[col])[0]
            for cells in grid.values() if col in cells
        ]
        if kinds and sum(1 for k in kinds if k == KIND_TEXT) / len(kinds) >= 0.5:
            text_cols.add(col)

    # --- emit rows ----------------------------------------------------------
    for price_date in sorted(grid):
        for col, value in grid[price_date].items():
            kind, low, high, text, status = classify_value(value)
            if col in text_cols and kind == KIND_STATUS and status:
                # A MONTH column holding an unrecognised label is still a label.
                kind, text, status = KIND_TEXT, status, None
            result.rows.append(ParsedRow(
                sheet_name=ws.title,
                series_name=series_by_col[col],
                column_index=col,
                price_date=price_date,
                raw_value=_clean(value),
                price_value=None if low is None else (low + high) / 2,
                price_low=low,
                price_high=high,
                text_value=text,
                status_flag=status,
                kind=kind,
            ))

    if min_d and max_d:
        result.date_range = (min_d, max_d)
    return result


def sanitize_identifier(name: str) -> str:
    """Turn a header string (or sheet name) into a valid unquoted Snowflake
    identifier, e.g. "PALM OIL | CRUDE" -> "PALM_OIL_CRUDE", "MISC 1" ->
    "MISC_1"."""
    cleaned = re.sub(r"[^A-Za-z0-9]+", "_", name).strip("_").upper()
    if not cleaned:
        cleaned = "COLUMN"
    if cleaned[0].isdigit():
        cleaned = f"C_{cleaned}"
    return cleaned


def pivot_sheet_wide(sheet: SheetParseResult):
    """Reshape a sheet's long rows into one row per PRICE_DATE, matching the
    source Excel tab: a DAY_TYPE column, then per price series either

      <SERIES>_LOW and <SERIES>_HIGH   (FLOAT; equal for a single quote), or
      <SERIES>                         (VARCHAR, for a futures MONTH column).

    DAY_TYPE is set only when EVERY series on that date shares the same
    non-numeric status token (e.g. the whole sheet says SUNDAY) -- a single
    series being NA for one date does not count as a sheet-wide day type.

    Columns are keyed by COLUMN_INDEX, not by the raw series-name text, since
    two columns can end up with the same reconstructed header text (a futures
    block repeats CLOSE/CHANGE per contract) -- COLUMN_INDEX is always unique.
    """
    import pandas as pd

    col_order: list[int] = []
    name_by_index: dict[int, str] = {}
    text_index: set = set()
    for row in sheet.rows:
        if row.column_index not in name_by_index:
            name_by_index[row.column_index] = row.series_name
            col_order.append(row.column_index)
        if row.kind == KIND_TEXT:
            text_index.add(row.column_index)

    seen: dict[str, int] = {}
    sanitized_by_index: dict[int, str] = {}
    for idx in col_order:
        base = sanitize_identifier(name_by_index[idx])
        n = seen.get(base, 0)
        sanitized_by_index[idx] = base if n == 0 else f"{base}_{n + 1}"
        seen[base] = n + 1

    by_date: dict[datetime.date, dict[int, ParsedRow]] = {}
    for row in sheet.rows:
        by_date.setdefault(row.price_date, {})[row.column_index] = row

    columns = ["PRICE_DATE", "DAY_TYPE"]
    types = {"PRICE_DATE": "DATE", "DAY_TYPE": "VARCHAR"}
    for idx in col_order:
        base = sanitized_by_index[idx]
        if idx in text_index:
            columns.append(base)
            types[base] = "VARCHAR"
        else:
            columns.extend([f"{base}_LOW", f"{base}_HIGH"])
            types[f"{base}_LOW"] = "FLOAT"
            types[f"{base}_HIGH"] = "FLOAT"

    records = []
    for price_date in sorted(by_date):
        cells = by_date[price_date]
        record = {"PRICE_DATE": price_date}
        statuses = []
        any_present = False
        any_value = False
        for idx in col_order:
            base = sanitized_by_index[idx]
            cell = cells.get(idx)
            if idx in text_index:
                record[base] = cell.text_value if cell is not None else None
                if cell is not None and cell.text_value:
                    any_present = True
                continue
            if cell is None:
                record[f"{base}_LOW"] = record[f"{base}_HIGH"] = None
                continue
            any_present = True
            if cell.price_low is not None:
                record[f"{base}_LOW"] = cell.price_low
                record[f"{base}_HIGH"] = cell.price_high
                any_value = True
            else:
                record[f"{base}_LOW"] = record[f"{base}_HIGH"] = None
                if cell.status_flag:
                    statuses.append(cell.status_flag)
        record["DAY_TYPE"] = (
            statuses[0]
            if any_present and not any_value and statuses and len(set(statuses)) == 1
            else None
        )
        records.append(record)

    sheet.wide_column_types = types
    return pd.DataFrame.from_records(records, columns=columns)


def parse_workbook(file_path_or_bytes, file_name: str) -> WorkbookParseResult:
    wb = openpyxl.load_workbook(file_path_or_bytes, data_only=True, read_only=False)
    result = WorkbookParseResult(file_name=file_name)
    for sheet_name in wb.sheetnames:
        if sheet_name.upper() in SKIP_SHEETS:
            continue
        ws = wb[sheet_name]
        result.sheets.append(parse_sheet(ws))
    return result


if __name__ == "__main__":
    import sys
    for path in sys.argv[1:]:
        wb_result = parse_workbook(path, path.split("/")[-1])
        print("=" * 100)
        print(wb_result.file_name, "-> total rows:", wb_result.total_rows)
        for s in wb_result.sheets:
            print(f"  {s.sheet_name:26s} series={len(s.series):3d} rows={len(s.rows):5d} "
                  f"date_range={s.date_range} warnings={s.warnings}")
